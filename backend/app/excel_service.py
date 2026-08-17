from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


REQUIRED_COLUMNS = ["order_id", "customer_name", "address", "delivery_time"]
HEADER_SYNONYMS = {
    "order_id": ["order_id", "order id", "order no", "order number", "si no", "si. no", "sr no", "sl no", "sl.no", "slno", "serial", "s.no", "no", "no.", "id", "order"],
    "customer_name": ["customer_name", "customer name", "customer", "name", "client", "recipient"],
    "address": ["address", "contact address", "delivery address", "location", "location address", "delivery location", "destination", "area", "address detail", "full address"],
    "delivery_time": ["delivery_time", "delivery time", "time", "slot", "delivery slot", "delivery window", "timing"],
    "lat": ["lat", "latitude"],
    "lng": ["lng", "lng.", "long", "longitude"],
}
MONTH_KEYWORDS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def normalize_header_cell(value: Optional[str]) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""

    normalized = text.replace("\n", " ").replace("_", " ").replace(".", " ").strip()
    normalized = " ".join(normalized.split())

    for canonical, synonyms in HEADER_SYNONYMS.items():
        for synonym in synonyms:
            if normalized == synonym:
                return canonical

    if any(month in normalized for month in MONTH_KEYWORDS) and any(char.isdigit() for char in normalized):
        return "delivery_time"

    if "time" in normalized or "slot" in normalized:
        return "delivery_time"

    return ""


def find_header_row(rows: List[List[Optional[object]]]) -> Optional[int]:
    for index, row in enumerate(rows[:20]):
        normalized = [normalize_header_cell(str(value).strip() if value is not None else "") for value in row]
        if all(column in normalized for column in REQUIRED_COLUMNS):
            return index
    return None


def validate_excel_file(file_path: str) -> Dict[str, object]:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return {"is_valid": False, "errors": ["Excel file is empty"], "total_orders": 0}

        header_index = find_header_row(rows)
        if header_index is None:
            return {
                "is_valid": False,
                "errors": ["Missing required columns: order_id, customer_name, address, delivery_time"],
                "total_orders": 0,
            }

        header_row = rows[header_index]
        raw_headers = [str(value).strip() if value is not None else "" for value in header_row]
        headers = [normalize_header_cell(value) for value in raw_headers]

        # The exact left-to-right column layout of the uploaded sheet - every
        # column, in its original position, under its original header text.
        # Columns we don't recognize (contact number, amount, payment mode,
        # remarks, box counts, whatever else a business tracks alongside its
        # orders) used to be silently dropped on upload - as was a second
        # column that canonicalizes to one we already filled (e.g. a
        # "LOCATION" column alongside "CONTACT ADDRESS" - both map to
        # `address`, so LOCATION lost every time). Both are kept now.
        # Each entry is {"label": <original header text>, "field": <one of
        # REQUIRED_COLUMNS, or None for an extra column read from
        # extra_fields[label]>} - this is what lets "Download all" rebuild
        # the upload's exact column order and wording, not a fixed layout
        # this app made up.
        column_order: List[Dict[str, Optional[str]]] = []
        seen_canonical_columns = set()
        seen_extra_headers = set()
        for column_name, raw_header in zip(headers, raw_headers):
            if column_name:
                if column_name in seen_canonical_columns:
                    if raw_header and raw_header not in seen_extra_headers:
                        seen_extra_headers.add(raw_header)
                        column_order.append({"label": raw_header, "field": None})
                else:
                    seen_canonical_columns.add(column_name)
                    column_order.append({"label": raw_header or column_name, "field": column_name})
                continue
            if raw_header and raw_header not in seen_extra_headers:
                seen_extra_headers.add(raw_header)
                column_order.append({"label": raw_header, "field": None})

        data_rows = rows[header_index + 1 :]
        valid_rows = []
        errors = []

        for row_offset, row in enumerate(data_rows, start=header_index + 2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            row_data = {}
            extra_fields: Dict[str, object] = {}
            for column_name, raw_header, cell_value in zip(headers, raw_headers, row):
                extra_value = cell_value.strip() if isinstance(cell_value, str) else cell_value
                if column_name:
                    if column_name in row_data:
                        # A second column also canonicalizes to a field
                        # that's already filled - e.g. a separate "LOCATION"
                        # column alongside "CONTACT ADDRESS" (both map to
                        # `address`). Keep the first as the canonical value,
                        # but don't silently lose the second - it goes into
                        # extra_fields under its own original header instead.
                        if raw_header and raw_header not in extra_fields:
                            extra_fields[raw_header] = extra_value
                        continue
                    row_data[column_name] = str(cell_value).strip() if cell_value is not None else ""
                elif raw_header and raw_header not in extra_fields:
                    extra_fields[raw_header] = extra_value
            row_data["extra_fields"] = extra_fields

            if not row_data.get("order_id"):
                errors.append(f"Row {row_offset}: order_id is missing")
                continue
            if not row_data.get("customer_name"):
                errors.append(f"Row {row_offset}: customer_name is missing")
                continue
            if not row_data.get("address"):
                errors.append(f"Row {row_offset}: address is missing")
                continue
            if not row_data.get("delivery_time"):
                errors.append(f"Row {row_offset}: delivery_time is missing")
                continue

            # Convert lat / lng if present in excel
            if row_data.get("lat"):
                try:
                    row_data["lat"] = float(row_data["lat"])
                except ValueError:
                    row_data["lat"] = None
            if row_data.get("lng"):
                try:
                    row_data["lng"] = float(row_data["lng"])
                except ValueError:
                    row_data["lng"] = None

            valid_rows.append(row_data)

        return {
            "is_valid": not errors,
            "errors": errors,
            "total_orders": len(valid_rows),
            "orders": valid_rows,
            "column_order": column_order,
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass
