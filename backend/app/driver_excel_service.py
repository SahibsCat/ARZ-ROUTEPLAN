"""Server-side per-route Excel export for the driver app. The admin panel
already builds a richer, styled version of this client-side (App.jsx's
addRouteWorksheet, via exceljs, in the browser) - that path stays exactly
as it is. This is a separate, plainer generator because the driver app
(a phone, not a browser with that JS bundle loaded) needs the server to
produce the file instead. Same columns, same source data
(crud.route_summary), deliberately not sharing code across languages."""

from io import BytesIO
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="FF334155", end_color="FF334155", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def _phone_from_extra_fields(order: Dict[str, object]) -> str:
    extra = order.get("extra_fields") or {}
    for label, value in extra.items():
        if str(label).strip().lower() in ("phone", "contact number", "contact", "mobile", "mobile number", "phone number"):
            if value:
                return str(value).strip()
    return ""


def build_route_workbook(route_summary: Dict[str, object]) -> BytesIO:
    wb = Workbook()
    sheet = wb.active
    sheet.title = str(route_summary.get("route_name") or "Route")[:31]

    headers = [
        "Stop #", "Order ID", "Customer", "Phone", "Address", "Location",
        "Delivery Slot", "ETA", "Status", "Delivered", "Route", "Vehicle", "Latitude", "Longitude", "Google Maps",
    ]
    for col, label in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    vehicle_label = "Car" if route_summary.get("vehicle_type") == "car" else "Bike"
    for idx, order in enumerate(route_summary.get("orders", []), start=1):
        row = idx + 1
        values = [
            idx,
            order.get("order_id"),
            order.get("customer_name"),
            _phone_from_extra_fields(order) or "—",
            order.get("address"),
            order.get("area") or "—",
            order.get("delivery_time"),
            order.get("eta") or "—",
            "LATE" if order.get("is_late") else "On time",
            "Delivered" if order.get("is_delivered") else "Pending",
            route_summary.get("route_name"),
            vehicle_label,
            order.get("lat") if order.get("lat") is not None else "—",
            order.get("lng") if order.get("lng") is not None else "—",
            order.get("map_link") or "—",
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row, column=col, value=value)

    widths = [8, 12, 20, 14, 38, 16, 14, 10, 10, 11, 12, 10, 12, 12, 40]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
