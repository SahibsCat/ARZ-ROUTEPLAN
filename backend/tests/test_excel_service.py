from pathlib import Path

from openpyxl import Workbook

from app.excel_service import validate_excel_file


def test_validate_excel_file_detects_missing_columns(tmp_path):
    file_path = tmp_path / "orders.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["order_id", "customer_name", "delivery_time"])
    sheet.append(["1001", "Alice", "10:00"])
    workbook.save(file_path)

    result = validate_excel_file(str(file_path))

    assert result["is_valid"] is False
    assert result["total_orders"] == 0
    assert "Missing required columns" in result["errors"][0]


def test_validate_excel_file_accepts_valid_rows(tmp_path):
    file_path = tmp_path / "orders.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["order_id", "customer_name", "address", "delivery_time"])
    sheet.append(["1001", "Alice", "123 Main St", "10:00"])
    sheet.append(["1002", "Bob", "456 Side Rd", "11:00"])
    workbook.save(file_path)

    result = validate_excel_file(str(file_path))

    assert result["is_valid"] is True
    assert result["total_orders"] == 2
    assert len(result["orders"]) == 2


def test_validate_excel_file_accepts_sample_header_synonyms(tmp_path):
    file_path = tmp_path / "orders.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['', '', 'SI. NO', 'CUSTOMER NAME', 'CONTACT NUMBER', 'CONTACT ADDRESS', 'LOCATION', '26th July2026', 'AMOUNT', 'PAYMENT MODE', 'REMARKS', 'EXTRA', 'TOTAL BOX FF5', 'TOTAL BOX FF2', 'WEEKEND FF5'])
    sheet.append(['', '', '1', 'udaya', '98415 81815', 'BBCL Stanburry Villa 17\nOld No: 56, New No: 79, Anna Main Rd, Kolapakkam, Chennai, Kolapakkam, Tamil Nadu 600128', 'kolapakkam', '12pm sharp', '3100', 'gpay/23', 'paid', 'new chicken biryani ff5-01,chicken 65-01', '1', '', '', ''])
    workbook.save(file_path)

    result = validate_excel_file(str(file_path))

    assert result["is_valid"] is True
    assert result["total_orders"] == 1
    assert result["orders"][0]["order_id"] == "1"
    assert result["orders"][0]["customer_name"] == "udaya"
    assert "BBCL Stanburry Villa" in result["orders"][0]["address"]
    assert result["orders"][0]["delivery_time"] == "12pm sharp"


def test_validate_excel_file_drops_duplicate_order_id(tmp_path):
    file_path = tmp_path / "orders.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["order_id", "customer_name", "address", "delivery_time"])
    sheet.append(["17", "praveen", "Ceebros Blvd, Mettukuppam, Chennai", "12pm to 1pm"])
    sheet.append(["18", "kumar", "Some other street, Chennai", "1pm to 2pm"])
    # Same order_id as row 2 (a real duplicate row, not a coincidence) - used
    # to become a second identical stop on whatever route order 17 landed
    # on. Only the first occurrence should survive.
    sheet.append(["17", "praveen", "Ceebros Blvd, Mettukuppam, Chennai", "12pm to 1pm"])
    workbook.save(file_path)

    result = validate_excel_file(str(file_path))

    assert result["total_orders"] == 2
    order_ids = [o["order_id"] for o in result["orders"]]
    assert order_ids == ["17", "18"]
    assert any("duplicate order_id" in e for e in result["errors"])


def test_validate_excel_file_preserves_extra_business_columns(tmp_path):
    file_path = tmp_path / "orders.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['', '', 'SI. NO', 'CUSTOMER NAME', 'CONTACT NUMBER', 'CONTACT ADDRESS', 'LOCATION', '26th July2026', 'AMOUNT', 'PAYMENT MODE', 'REMARKS', 'EXTRA', 'TOTAL BOX FF5', 'TOTAL BOX FF2', 'WEEKEND FF5'])
    sheet.append(['', '', '1', 'udaya', '98415 81815', 'BBCL Stanburry Villa 17, Kolapakkam, Chennai', 'kolapakkam', '12pm sharp', 3100, 'gpay/23', 'paid', 'new chicken biryani ff5-01', 1, '', ''])
    workbook.save(file_path)

    result = validate_excel_file(str(file_path))

    assert result["is_valid"] is True
    order = result["orders"][0]
    # Required fields still populated as before.
    assert order["order_id"] == "1"
    assert order["address"] == "BBCL Stanburry Villa 17, Kolapakkam, Chennai"

    extra = order["extra_fields"]
    assert extra["CONTACT NUMBER"] == "98415 81815"
    # LOCATION also canonicalizes to `address`, but CONTACT ADDRESS already
    # claimed that slot - it must survive under its own header, not vanish.
    assert extra["LOCATION"] == "kolapakkam"
    assert extra["AMOUNT"] == 3100
    assert extra["PAYMENT MODE"] == "gpay/23"
    assert extra["REMARKS"] == "paid"
    assert extra["TOTAL BOX FF5"] == 1

    # column_order captures the exact original layout - every column, in
    # its original left-to-right position and header text, whether it's one
    # of the 4 required fields or an extra one - so an export can rebuild
    # the upload's exact column order and wording later.
    column_order = result["column_order"]
    labels = [c["label"] for c in column_order]
    assert labels == [
        "SI. NO", "CUSTOMER NAME", "CONTACT NUMBER", "CONTACT ADDRESS", "LOCATION",
        "26th July2026", "AMOUNT", "PAYMENT MODE", "REMARKS", "EXTRA",
        "TOTAL BOX FF5", "TOTAL BOX FF2", "WEEKEND FF5",
    ]
    fields = {c["label"]: c["field"] for c in column_order}
    assert fields["SI. NO"] == "order_id"
    assert fields["CONTACT ADDRESS"] == "address"
    assert fields["26th July2026"] == "delivery_time"
    assert fields["CONTACT NUMBER"] is None
    assert fields["LOCATION"] is None
    assert fields["AMOUNT"] is None
